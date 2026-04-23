"""
ARKASku - Aplikasi Tarik Data ARKAS ke Excel
===================================================
Versi: 2.1 - Hacker Premium Edition
Author: Operator SD Negeri Pasirhalang
"""

import os
import json
import sqlcipher3
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from datetime import datetime

app = Flask(__name__)

# === KONFIGURASI ===
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

def load_config():
    """Load konfigurasi dari config.json"""
    if os.path.exists(CONFIG_FILE):
        print(f"📂 Menggunakan config dari: {CONFIG_FILE}")
        with open(CONFIG_FILE, 'r') as f:
            config = json.load(f)
        return {
            'db_path': config['db_path'],
            'db_key': config['db_key'],
            'sekolah': config['sekolah'],
            'npsn': config['npsn']
        }
    
    # Fallback
    print("⚠️ CONFIG.TIDAK KETEMU! BUAT FILE config.json DI FOLDER PROJECT!")
    return {
        'db_path': r"arkas.db",
        'db_key': "K3md1kbudRIS3n4yan",
        'sekolah': "SEKOLAH ANDA",
        'npsn': "NPSN_ANDA"
    }

# Load konfigurasi
CONFIG = load_config()
DB_PATH = CONFIG['db_path']
DB_KEY = CONFIG['db_key']
SEKOLAH = CONFIG['sekolah']
NPSN = CONFIG['npsn']

# === DATABASE FUNCTIONS ===
def get_db():
    """Koneksi ke database ARKAS"""
    # 1. Coba Standar SQLite
    try:
        db = sqlcipher3.connect(DB_PATH)
        db.execute("SELECT name FROM sqlite_master WHERE type='table' LIMIT 1")
        print("✅ Terhubung (Mode: Standar SQLite)")
        return db
    except:
        pass
    
    # 2. Coba SQLCipher 4
    try:
        db = sqlcipher3.connect(DB_PATH)
        db.execute(f"PRAGMA key = '{DB_KEY}'")
        db.execute("PRAGMA cipher_compatibility = 4")
        db.execute("SELECT name FROM sqlite_master WHERE type='table' LIMIT 1")
        print("✅ Terhubung (Mode: SQLCipher 4)")
        return db
    except:
        pass
    
    # 3. Coba SQLCipher 3
    try:
        db = sqlcipher3.connect(DB_PATH)
        db.execute(f"PRAGMA key = '{DB_KEY}'")
        db.execute("PRAGMA cipher_compatibility = 3")
        db.execute("SELECT name FROM sqlite_master WHERE type='table' LIMIT 1")
        print("✅ Terhubung (Mode: SQLCipher 3)")
        return db
    except Exception as e:
        print(f"❌ Gagal koneksi database: {e}")
        raise e

def query_arkas(sql, params=None):
    """Eksekusi query ARKAS"""
    try:
        db = get_db()
        if params:
            cursor = db.execute(sql, params)
        else:
            cursor = db.execute(sql)
        rows = cursor.fetchall()
        db.close()
        return rows
    except Exception as e:
        print(f"Error query: {e}")
        return []

def get_all_tables():
    """Ambil semua nama tabel dari database"""
    rows = query_arkas("""
        SELECT name FROM sqlite_master 
        WHERE type='table' AND name NOT LIKE 'sqlite_%' 
        ORDER BY name
    """)
    return [r[0] for r in rows]

def get_list_anggaran():
    """Ambil daftar tahun anggaran yang sah (soft_delete=0)"""
    rows = query_arkas("""
        SELECT id_anggaran, tahun_anggaran, jumlah 
        FROM anggaran 
        WHERE soft_delete = 0
        ORDER BY tahun_anggaran DESC, create_date DESC
    """)
    # Ambil hanya yang terbaru per tahun jika ada multiple approved (jarang tapi mungkin)
    seen = set()
    unique_rows = []
    for r in rows:
        if r[1] not in seen:
            seen.add(r[1])
            unique_rows.append(r)
    return unique_rows

def get_anggaran_terbaru():
    """Ambil anggaran terbaru yang sah"""
    rows = query_arkas("""
        SELECT id_anggaran, tahun_anggaran, jumlah 
        FROM anggaran 
        WHERE soft_delete = 0
        ORDER BY create_date DESC LIMIT 1
    """)
    return rows[0] if rows else None

def get_latest_id_anggaran(tahun):
    """Helper untuk mendapatkan ID anggaran sah terbaru untuk tahun tertentu"""
    if not tahun:
        return None
    rows = query_arkas("""
        SELECT id_anggaran FROM anggaran 
        WHERE tahun_anggaran = ? AND soft_delete = 0 
        ORDER BY create_date DESC LIMIT 1
    """, (tahun,))
    return rows[0][0] if rows else None

def get_rapbs(id_anggaran):
    """Ambil data RAPBS"""
    if not id_anggaran:
        return []
    return query_arkas(f"""
        SELECT kode_rekening, uraian, volume, satuan, harga_satuan, jumlah 
        FROM rapbs 
        WHERE id_anggaran = '{id_anggaran}'
        ORDER BY kode_rekening
    """)

def get_kas_umum(limit=None, tahun=None, bulan=None, tahapan=None):
    """Ambil data Kas Umum (filter soft_delete=0)"""
    id_anggaran = get_latest_id_anggaran(tahun)
    where_clause = " WHERE k.soft_delete = 0"
    
    if id_anggaran:
        where_clause += f" AND k.id_anggaran = '{id_anggaran}'"
    elif tahun:
        where_clause += f" AND CAST(strftime('%Y', k.tanggal_transaksi) AS INTEGER) = {tahun}"
        
    if bulan:
        where_clause += f" AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) = {bulan}"
    
    if tahapan:
        if tahapan == 1:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 1 AND 6"
        elif tahapan == 2:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 7 AND 12"
    
    # Query dioptimasi untuk menghindari timeout jika mungkin (JOIN ref_bku tetap dilakukan)
    return query_arkas(f"""
        SELECT 
            k.tanggal_transaksi,
            r.bku as status,
            k.kode_rekening,
            k.no_bukti,
            k.uraian,
            CASE 
                WHEN k.saldo > COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                THEN k.saldo - COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                ELSE 0
            END as pemasukkan,
            CASE 
                WHEN k.saldo < COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo)
                THEN COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo) - k.saldo
                ELSE 0
            END as pengeluaran,
            k.saldo
        FROM kas_umum k
        LEFT JOIN ref_bku r ON k.id_ref_bku = r.id_ref_bku
        {where_clause}
        ORDER BY k.tanggal_transaksi ASC, k.create_date ASC
    """)

def get_kas_bank(tahun=None, bulan=None, tahapan=None):
    """Ambil data Kas Pembantu Bank (id_ref_bku IN (2, 8), filter soft_delete=0)"""
    id_anggaran = get_latest_id_anggaran(tahun)
    where_clause = " WHERE k.id_ref_bku IN (2, 8) AND k.soft_delete = 0"
    
    if id_anggaran:
        where_clause += f" AND k.id_anggaran = '{id_anggaran}'"
    elif tahun:
        where_clause += f" AND CAST(strftime('%Y', k.tanggal_transaksi) AS INTEGER) = {tahun}"
        
    if bulan:
        where_clause += f" AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) = {bulan}"
    
    if tahapan:
        if tahapan == 1:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 1 AND 6"
        elif tahapan == 2:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 7 AND 12"
    
    return query_arkas(f"""
        SELECT 
            k.tanggal_transaksi,
            r.bku as status,
            k.kode_rekening,
            k.no_bukti,
            k.uraian,
            CASE 
                WHEN k.saldo > COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                THEN k.saldo - COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                ELSE 0
            END as pemasukkan,
            CASE 
                WHEN k.saldo < COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo)
                THEN COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo) - k.saldo
                ELSE 0
            END as pengeluaran,
            k.saldo
        FROM kas_umum k
        LEFT JOIN ref_bku r ON k.id_ref_bku = r.id_ref_bku
        {where_clause}
        ORDER BY k.tanggal_transaksi ASC, k.create_date ASC
    """)

def get_kas_pajak(tahun=None, bulan=None, tahapan=None):
    """Ambil data Kas Pembantu Pajak (id_ref_bku IN (7, 10), filter soft_delete=0)"""
    id_anggaran = get_latest_id_anggaran(tahun)
    where_clause = " WHERE k.id_ref_bku IN (7, 10) AND k.soft_delete = 0"
    
    if id_anggaran:
        where_clause += f" AND k.id_anggaran = '{id_anggaran}'"
    elif tahun:
        where_clause += f" AND CAST(strftime('%Y', k.tanggal_transaksi) AS INTEGER) = {tahun}"
        
    if bulan:
        where_clause += f" AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) = {bulan}"
    
    if tahapan:
        if tahapan == 1:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 1 AND 6"
        elif tahapan == 2:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 7 AND 12"
    
    return query_arkas(f"""
        SELECT 
            k.tanggal_transaksi,
            r.bku as status,
            k.kode_rekening,
            k.no_bukti,
            k.uraian,
            CASE 
                WHEN k.saldo > COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                THEN k.saldo - COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                ELSE 0
            END as pemasukkan,
            CASE 
                WHEN k.saldo < COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo)
                THEN COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo) - k.saldo
                ELSE 0
            END as pengeluaran,
            k.saldo
        FROM kas_umum k
        LEFT JOIN ref_bku r ON k.id_ref_bku = r.id_ref_bku
        {where_clause}
        ORDER BY k.tanggal_transaksi ASC, k.create_date ASC
    """)

def get_kas_tunai(tahun=None, bulan=None, tahapan=None):
    """Ambil data Kas Pembantu Tunai (id_ref_bku IN (3, 5, 9), filter soft_delete=0)"""
    id_anggaran = get_latest_id_anggaran(tahun)
    where_clause = " WHERE k.id_ref_bku IN (3, 5, 9) AND k.soft_delete = 0"
    
    if id_anggaran:
        where_clause += f" AND k.id_anggaran = '{id_anggaran}'"
    elif tahun:
        where_clause += f" AND CAST(strftime('%Y', k.tanggal_transaksi) AS INTEGER) = {tahun}"
        
    if bulan:
        where_clause += f" AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) = {bulan}"
    
    if tahapan:
        if tahapan == 1:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 1 AND 6"
        elif tahapan == 2:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 7 AND 12"
    
    return query_arkas(f"""
        SELECT 
            k.tanggal_transaksi,
            r.bku as status,
            k.kode_rekening,
            k.no_bukti,
            k.uraian,
            CASE 
                WHEN k.saldo > COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                THEN k.saldo - COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                ELSE 0
            END as pemasukkan,
            CASE 
                WHEN k.saldo < COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo)
                THEN COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo) - k.saldo
                ELSE 0
            END as pengeluaran,
            k.saldo
        FROM kas_umum k
        LEFT JOIN ref_bku r ON k.id_ref_bku = r.id_ref_bku
        {where_clause}
        ORDER BY k.tanggal_transaksi ASC, k.create_date ASC
    """)
    if tahapan:
        if tahapan == 1:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 1 AND 6"
        elif tahapan == 2:
            where_clause += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 7 AND 12"
    
    return query_arkas(f"""
        SELECT 
            k.tanggal_transaksi,
            r.bku as status,
            k.kode_rekening,
            k.no_bukti,
            k.uraian,
            CASE 
                WHEN k.saldo > COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                THEN k.saldo - COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), 0)
                ELSE 0
            END as pemasukkan,
            CASE 
                WHEN k.saldo < COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo)
                THEN COALESCE((SELECT k2.saldo FROM kas_umum k2 
                    WHERE k2.tanggal_transaksi < k.tanggal_transaksi 
                    ORDER BY k2.tanggal_transaksi DESC LIMIT 1), k.saldo) - k.saldo
                ELSE 0
            END as pengeluaran,
            k.saldo
        FROM kas_umum k
        LEFT JOIN ref_bku r ON k.id_ref_bku = r.id_ref_bku
        {where_clause}
        ORDER BY k.tanggal_transaksi ASC
    """)

def get_kertas_kerja(tahun=None, bulan=None, tahapan=None):
    """Ambil data Kertas Kerja dari rapbs - sesuai format PDF (filter id_anggaran sah)"""
    id_anggaran = get_latest_id_anggaran(tahun)
    if not id_anggaran:
        return []
        
    where_clause = f" WHERE id_anggaran = '{id_anggaran}'"
    # Handle semester: tahapan=1 (SEM_1: bln 1-6), tahapan=2 (SEM_2: bln 7-12)
    if tahapan:
        if tahapan == 1:
            where_clause += " AND SUBSTR(kode_rekening, 6, 2) IN ('01','02','03','04','05','06')"
        elif tahapan == 2:
            where_clause += " AND SUBSTR(kode_rekening, 6, 2) IN ('07','08','09','10','11','12')"
    
    return query_arkas(f"""
        SELECT kode_rekening, uraian, volume, satuan, jumlah
        FROM rapbs
        {where_clause}
        GROUP BY kode_rekening
        ORDER BY kode_rekening
    """)

def get_rkas(tahun=None, bulan=None, tahapan=None):
    """Ambil data RKAS dari rapbs - sesuai format PDF (filter id_anggaran sah)"""
    id_anggaran = get_latest_id_anggaran(tahun)
    if not id_anggaran:
        return []

    where_clause = f" WHERE id_anggaran = '{id_anggaran}'"
    if tahapan:
        if tahapan == 1:
            where_clause += " AND SUBSTR(kode_rekening, 6, 2) IN ('01','02','03','04','05','06')"
        elif tahapan == 2:
            where_clause += " AND SUBSTR(kode_rekening, 6, 2) IN ('07','08','09','10','11','12')"
    
    return query_arkas(f"""
        SELECT kode_rekening, uraian, volume, satuan, jumlah
        FROM rapbs
        {where_clause}
        GROUP BY kode_rekening
        ORDER BY kode_rekening
    """)

def get_realisasi(tahun=None, bulan=None, tahapan=None):
    """Ambil data Realisasi - anggaran + kas_umum (format PDF: anggaran vs realisasi)"""
    id_anggaran = get_latest_id_anggaran(tahun)
    if not id_anggaran:
        return []
    
    # Filter tambahan untuk kas_umum berdasarkan bulan/tahapan jika diperlukan
    sub_where = f"WHERE k.id_anggaran = '{id_anggaran}' AND k.soft_delete = 0"
    if tahapan:
        if tahapan == 1:
            sub_where += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 1 AND 6"
        elif tahapan == 2:
            sub_where += " AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) BETWEEN 7 AND 12"
    if bulan:
        sub_where += f" AND CAST(strftime('%m', k.tanggal_transaksi) AS INTEGER) = {bulan}"
    
    return query_arkas(f"""
        SELECT r.kode_rekening, r.uraian, 
               r.jumlah as anggaran, 
               COALESCE((SELECT SUM(k.saldo) FROM kas_umum k {sub_where} AND k.kode_rekening = r.kode_rekening), 0) as realisasi,
               r.jumlah - COALESCE((SELECT SUM(k.saldo) FROM kas_umum k {sub_where} AND k.kode_rekening = r.kode_rekening), 0) as selisih
        FROM rapbs r
        WHERE r.id_anggaran = '{id_anggaran}'
        ORDER BY r.kode_rekening
    """)

def get_realisasi_barang_habis(tahun=None, bulan=None, tahapan=None):
    """Ambil data BHP - filter id_anggaran sah (prefix 5.1.02.01 = supplies) - Tanpa Konsumsi - Volume RAPBS"""
    id_anggaran = get_latest_id_anggaran(tahun)
    if not id_anggaran:
        return []
        
    kode_filter = " AND SUBSTR(kode_rekening, 1, 9) = '5.1.02.01'"
    tahap_filter = ""
    if tahapan:
        if tahapan == 1:
            tahap_filter = " AND SUBSTR(r.kode_rekening, 6, 2) IN ('01','02','03','04','05','06')"
        elif tahapan == 2:
            tahap_filter = " AND SUBSTR(r.kode_rekening, 6, 2) IN ('07','08','09','10','11','12')"
    
    return query_arkas(f"""
        SELECT r.kode_rekening, r.uraian, r.volume, r.satuan, r.harga_satuan, r.jumlah
        FROM rapbs r
        WHERE r.id_anggaran = '{id_anggaran}' {kode_filter}{tahap_filter}
          AND LOWER(r.uraian) NOT LIKE '%konsumsi%'
          AND LOWER(r.uraian) NOT LIKE '%snack%'
          AND LOWER(r.uraian) NOT LIKE '%makan%'
          AND LOWER(r.uraian) NOT LIKE '%minum%'
          AND r.kode_rekening NOT LIKE '%.0052'
          AND r.kode_rekening NOT LIKE '%.0055'
          AND r.kode_rekening NOT LIKE '%.0054'
        ORDER BY r.kode_rekening, r.volume
    """)

def get_realisasi_barang_modal(tahun=None, bulan=None, tahapan=None):
    """Ambil data Modal/Aset - filter id_anggaran sah (prefix 5.2 = modal) - Volume RAPBS"""
    id_anggaran = get_latest_id_anggaran(tahun)
    if not id_anggaran:
        return []
        
    kode_filter = " AND SUBSTR(r.kode_rekening, 1, 3) = '5.2'"
    tahap_filter = ""
    if tahapan:
        if tahapan == 1:
            tahap_filter = " AND SUBSTR(r.kode_rekening, 6, 2) IN ('01','02','03','04','05','06')"
        elif tahapan == 2:
            tahap_filter = " AND SUBSTR(r.kode_rekening, 6, 2) IN ('07','08','09','10','11','12')"
    
    return query_arkas(f"""
        SELECT r.kode_rekening, r.uraian, r.volume, r.satuan, r.harga_satuan, r.jumlah
        FROM rapbs r
        WHERE r.id_anggaran = '{id_anggaran}' {kode_filter}{tahap_filter}
        ORDER BY r.kode_rekening, r.volume
    """)

def get_buku_pembantu_objek(tahun=None, bulan=None, tahapan=None):
    """Ambil data Buku Pembantu Rincian Objek Belanja - sesuai format PDF (filter id_anggaran sah)"""
    id_anggaran = get_latest_id_anggaran(tahun)
    if not id_anggaran:
        return []

    where_clause = f" WHERE id_anggaran = '{id_anggaran}'"
    # Handle semester
    if tahapan:
        if tahapan == 1:
            where_clause += " AND SUBSTR(kode_rekening, 6, 2) IN ('01','02','03','04','05','06')"
        elif tahapan == 2:
            where_clause += " AND SUBSTR(kode_rekening, 6, 2) IN ('07','08','09','10','11','12')"
    
    return query_arkas(f"""
        SELECT kode_rekening, MAX(uraian) as uraian, SUM(volume) as volume, MAX(satuan) as satuan, SUM(jumlah) as jumlah
        FROM rapbs 
        {where_clause}
        GROUP BY kode_rekening
        ORDER BY kode_rekening
    """)

def get_laporan_bosp(tahun=None, semester=None):
    """Ambil data Laporan BOSP - ringkasan anggaran per tahun (filter soft_delete=0)"""
    id_anggaran = get_latest_id_anggaran(tahun)
    where_clause = " WHERE a.soft_delete = 0"
    if id_anggaran:
        where_clause += f" AND a.id_anggaran = '{id_anggaran}'"
    elif tahun:
        where_clause += f" AND a.tahun_anggaran = {tahun}"
    
    return query_arkas(f"""
        SELECT a.tahun_anggaran as tahun, a.jumlah as total_anggaran, 
               COALESCE((SELECT SUM(k.saldo) FROM kas_umum k WHERE k.id_anggaran = a.id_anggaran AND k.soft_delete = 0), 0) as total_realisasi,
               a.jumlah - COALESCE((SELECT SUM(k.saldo) FROM kas_umum k WHERE k.id_anggaran = a.id_anggaran AND k.soft_delete = 0), 0) as selisih,
               CASE WHEN a.is_approve = 1 THEN 'Approved' ELSE 'Pending' END as status
        FROM anggaran a
        {where_clause}
        ORDER BY a.tahun_anggaran DESC, a.create_date DESC
    """)

# === EXPORT FUNCTIONS ===
def format_rupiah(value):
    """Format number to Rupiah string"""
    if value is None or value == "":
        return ""
    try:
        num = float(value)
        if num == int(num):
            return f"Rp {int(num):,}".replace(",", ".")
        return f"Rp {num:,.2f}".replace(",", ".")
    except:
        return str(value)

def export_to_excel(data, headers, filename, title=None):
    """Export data ke Excel dengan format raport"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Title
    if title:
        ws.merge_cells('A1:' + ws.cell(row=1, column=len(headers)).coordinate)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        start_row = 3
    else:
        start_row = 1
    
    # Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_alignment = Alignment(horizontal='center')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Identify numeric columns (columns 6 = Pemasukkan, 7 = Pengeluaran, 8 = Saldo)
    numeric_cols = {6, 7, 8}
    
    # Data and footer sums
    total_pemasukkan = 0
    total_pengeluaran = 0
    total_saldo = 0
    
    for row_idx, row_data in enumerate(data, start_row + 1):
        for col, value in enumerate(row_data, 1):
            # Handle None values
            if value is None:
                value = ""
            # Handle date values
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d")
            # Format Rupiah for numeric columns
            elif col in numeric_cols and value not in ("", None):
                try:
                    num = float(value)
                    if col == 6:
                        total_pemasukkan += num
                    elif col == 7:
                        total_pengeluaran += num
                    elif col == 8:
                        total_saldo = num  # Last saldo
                    value = format_rupiah(num)
                except:
                    pass
            ws.cell(row=row_idx, column=col, value=value)
    
    # Footer row with totals
    footer_row = start_row + len(data) + 1
    ws.cell(row=footer_row, column=1, value="JUMLAH")
    ws.cell(row=footer_row, column=1).font = Font(bold=True)
    ws.merge_cells(f'A{footer_row}:E{footer_row}')
    ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal='right')
    
    # Total pemasukkan
    ws.cell(row=footer_row, column=6, value=format_rupiah(total_pemasukkan))
    ws.cell(row=footer_row, column=6).font = Font(bold=True)
    
    # Total pengeluaran
    ws.cell(row=footer_row, column=7, value=format_rupiah(total_pengeluaran))
    ws.cell(row=footer_row, column=7).font = Font(bold=True)
    
    # Last saldo
    ws.cell(row=footer_row, column=8, value=format_rupiah(total_saldo))
    ws.cell(row=footer_row, column=8).font = Font(bold=True)
    
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        try:
            column = col[0].column_letter
        except AttributeError:
            continue
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# === ROUTES ===
@app.route('/')
def index():
    """Halaman utama"""
    tables = get_all_tables()
    anggaran_list = get_list_anggaran()
    ang = get_anggaran_terbaru()
    
    return render_template('index.html', 
                    sekolah=SEKOLAH,
                    npsn=NPSN,
                    tables=tables,
                    anggaran_list=anggaran_list,
                    anggaran_terbaru=ang)

@app.route('/api/info')
def api_info():
    """API info sekolah"""
    rows = query_arkas("""
        SELECT nama, npsn, alamat, kepsek, nip_kepsek, jumlah_siswa 
        FROM mst_sekolah LIMIT 1
    """)
    if rows:
        info = rows[0]
        return jsonify({
            'nama': info[0],
            'npsn': info[1],
            'alamat': info[2],
            'kepsek': info[3],
            'nip_kepsek': info[4],
            'jumlah_siswa': info[5]
        })
    return jsonify({'error': 'Data tidak ditemukan'})

@app.route('/api/tabel')
def api_tabel():
    """API list semua tabel"""
    tables = get_all_tables()
    return jsonify({'tables': tables})

@app.route('/api/anggaran')
def api_anggaran():
    """API list anggaran"""
    anggaran = get_list_anggaran()
    return jsonify({
        'anggaran': [{'id': a[0], 'tahun': a[1], 'jumlah': a[2]} for a in anggaran]
    })

# ========== BUKU KAS UMUM ==========
@app.route('/export/bku-tahunan')
def export_bku_tahunan():
    """Export Buku Kas Umum Tahunan"""
    tahun = request.args.get('tahun')
    data = get_kas_umum(tahun=tahun)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    output = export_to_excel(
        data, headers, 
        f'Buku_Kas_Umum_Tahunan_{tahun or "Semua"}.xlsx',
        f'BUKU KAS UMUM TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BKU_Tahunan_{tahun or "Semua"}.xlsx')

@app.route('/export/bku-bulanan')
def export_bku_bulanan():
    """Export Buku Kas Umum Bulanan"""
    tahun = request.args.get('tahun')
    bulan = request.args.get('bulan', type=int)
    data = get_kas_umum(tahun=tahun, bulan=bulan)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    nama_bulan = bulan_nama[bulan-1] if bulan else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Buku_Kas_Umum_Bulanan_{tahun}_{nama_bulan}.xlsx',
        f'BUKU KAS UMUM BULAN {nama_bulan} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BKU_Bulanan_{tahun}_{nama_bulan}.xlsx')

@app.route('/export/bku-semester')
def export_bku_semester():
    """Export Buku Kas Umum Semester"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan', type=int)
    data = get_kas_umum(tahun=tahun, tahapan=tahapan)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    semester_nama = "SEM_1" if tahapan == 1 else "SEM_2" if tahapan == 2 else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Buku_Kas_Umum_{semester_nama}_{tahun}.xlsx',
        f'BUKU KAS UMUM {semester_nama} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BKU_{semester_nama}_{tahun}.xlsx')

# ========== BUKU KAS PEMBANTU BANK ==========
@app.route('/export/kas-bank-bulanan')
def export_kas_bank_bulanan():
    """Export Buku Kas Pembantu Bank Bulanan"""
    tahun = request.args.get('tahun')
    bulan = request.args.get('bulan', type=int)
    data = get_kas_bank(tahun=tahun, bulan=bulan)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    nama_bulan = bulan_nama[bulan-1] if bulan else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Kas_Pembantu_Bank_{tahun}_{nama_bulan}.xlsx',
        f'BUKU KAS PEMBANTU BANK BULAN {nama_bulan} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Kas_Bank_Bulanan_{tahun}_{nama_bulan}.xlsx')

@app.route('/export/kas-bank-tahunan')
def export_kas_bank_tahunan():
    """Export Buku Kas Pembantu Bank Tahunan"""
    tahun = request.args.get('tahun')
    data = get_kas_bank(tahun=tahun)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    output = export_to_excel(
        data, headers,
        f'Kas_Pembantu_Bank_Tahunan_{tahun}.xlsx',
        f'BUKU KAS PEMBANTU BANK TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Kas_Bank_Tahunan_{tahun}.xlsx')

@app.route('/export/kas-bank-semester')
def export_kas_bank_semester():
    """Export Buku Kas Pembantu Bank Semester"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan', type=int)
    data = get_kas_bank(tahun=tahun, tahapan=tahapan)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    semester_nama = "SEM_1" if tahapan == 1 else "SEM_2" if tahapan == 2 else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Kas_Pembantu_Bank_{semester_nama}_{tahun}.xlsx',
        f'BUKU KAS PEMBANTU BANK {semester_nama} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Kas_Bank_{semester_nama}_{tahun}.xlsx')

# ========== BUKU KAS PEMBANTU PAJAK ==========
@app.route('/export/kas-pajak-bulanan')
def export_kas_pajak_bulanan():
    """Export Buku Kas Pembantu Pajak Bulanan"""
    tahun = request.args.get('tahun')
    bulan = request.args.get('bulan', type=int)
    data = get_kas_pajak(tahun=tahun, bulan=bulan)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    nama_bulan = bulan_nama[bulan-1] if bulan else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Kas_Pembantu_Pajak_{tahun}_{nama_bulan}.xlsx',
        f'BUKU KAS PEMBANTU PAJAK BULAN {nama_bulan} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Kas_Pajak_Bulanan_{tahun}_{nama_bulan}.xlsx')

@app.route('/export/kas-pajak-tahunan')
def export_kas_pajak_tahunan():
    """Export Buku Kas Pembantu Pajak Tahunan"""
    tahun = request.args.get('tahun')
    data = get_kas_pajak(tahun=tahun)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    output = export_to_excel(
        data, headers,
        f'Kas_Pembantu_Pajak_Tahunan_{tahun}.xlsx',
        f'BUKU KAS PEMBANTU PAJAK TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Kas_Pajak_Tahunan_{tahun}.xlsx')

@app.route('/export/kas-pajak-semester')
def export_kas_pajak_semester():
    """Export Buku Kas Pembantu Pajak Semester"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan', type=int)
    data = get_kas_pajak(tahun=tahun, tahapan=tahapan)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    semester_nama = "SEM_1" if tahapan == 1 else "SEM_2" if tahapan == 2 else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Kas_Pembantu_Pajak_{semester_nama}_{tahun}.xlsx',
        f'BUKU KAS PEMBANTU PAJAK {semester_nama} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Kas_Pajak_{semester_nama}_{tahun}.xlsx')

# ========== BUKU KAS PEMBANTU TUNAI ==========
@app.route('/export/kas-tunai-bulanan')
def export_kas_tunai_bulanan():
    """Export Buku Kas Pembantu Tunai Bulanan"""
    tahun = request.args.get('tahun')
    bulan = request.args.get('bulan', type=int)
    data = get_kas_tunai(tahun=tahun, bulan=bulan)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    nama_bulan = bulan_nama[bulan-1] if bulan else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Kas_Pembantu_Tunai_{tahun}_{nama_bulan}.xlsx',
        f'BUKU KAS PEMBANTU TUNAI BULAN {nama_bulan} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Kas_Tunai_Bulanan_{tahun}_{nama_bulan}.xlsx')

@app.route('/export/kas-tunai-tahunan')
def export_kas_tunai_tahunan():
    """Export Buku Kas Pembantu Tunai Tahunan"""
    tahun = request.args.get('tahun')
    data = get_kas_tunai(tahun=tahun)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    output = export_to_excel(
        data, headers,
        f'Kas_Pembantu_Tunai_Tahunan_{tahun}.xlsx',
        f'BUKU KAS PEMBANTU TUNAI TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Kas_Tunai_Tahunan_{tahun}.xlsx')

@app.route('/export/kas-tunai-semester')
def export_kas_tunai_semester():
    """Export Buku Kas Pembantu Tunai Semester"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan', type=int)
    data = get_kas_tunai(tahun=tahun, tahapan=tahapan)
    headers = ['Tanggal', 'Status', 'Kode Rekening', 'No. Bukti', 'Uraian', 'Pemasukkan', 'Pengeluaran', 'Saldo']
    
    semester_nama = "SEM_1" if tahapan == 1 else "SEM_2" if tahapan == 2 else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Kas_Pembantu_Tunai_{semester_nama}_{tahun}.xlsx',
        f'BUKU KAS PEMBANTU TUNAI {semester_nama} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Kas_Tunai_{semester_nama}_{tahun}.xlsx')

# ========== RINCIAN KERTAS KERJA ==========
@app.route('/export/kk-tahunan')
def export_kk_tahunan():
    """Export Rincian Kertas Kerja Tahunan"""
    tahun = request.args.get('tahun')
    data = get_kertas_kerja(tahun=tahun)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    output = export_to_excel(
        data, headers,
        f'Kertas_Kerja_Tahunan_{tahun}.xlsx',
        f'RINCIAN KERTAS KERJA TAHUNAN TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'KK_Tahunan_{tahun}.xlsx')

@app.route('/export/kk-tahapan')
def export_kk_tahapan():
    """Export Rincian Kertas Kerja Tahapan"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan')
    data = get_kertas_kerja(tahun=tahun, tahapan=tahapan)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    output = export_to_excel(
        data, headers,
        f'Kertas_Kerja_Tahapan_{tahapan}_{tahun}.xlsx',
        f'RINCIAN KERTAS KERJA TAHAPAN {tahapan or "Semua"} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'KK_Tahapan_{tahapan}_{tahun}.xlsx')

@app.route('/export/kk-bulanan')
def export_kk_bulanan():
    """Export Rincian Kertas Kerja Bulanan"""
    tahun = request.args.get('tahun')
    # Untuk bulanan bisa ditambahkan filter bulan
    data = get_kertas_kerja(tahun=tahun)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    output = export_to_excel(
        data, headers,
        f'Kertas_Kerja_Bulanan_{tahun}.xlsx',
        f'RINCIAN KERTAS KERJA BULANAN TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'KK_Bulanan_{tahun}.xlsx')

# ========== RINCIAN RKAS ==========
@app.route('/export/rkas-tahunan')
def export_rkas_tahunan():
    """Export Rincian RKAS Tahunan"""
    tahun = request.args.get('tahun')
    data = get_rkas(tahun=tahun)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    output = export_to_excel(
        data, headers,
        f'RKAS_Tahunan_{tahun}.xlsx',
        f'RINCIAN RKAS TAHUNAN TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'RKAS_Tahunan_{tahun}.xlsx')

@app.route('/export/rkas-tahapan')
def export_rkas_tahapan():
    """Export Rincian RKAS Tahapan"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan')
    data = get_rkas(tahun=tahun, tahapan=tahapan)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    output = export_to_excel(
        data, headers,
        f'RKAS_Tahapan_{tahapan}_{tahun}.xlsx',
        f'RINCIAN RKAS TAHAPAN {tahapan or "Semua"} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'RKAS_Tahapan_{tahapan}_{tahun}.xlsx')

# ========== REKAPITULASI REALISASI ==========
@app.route('/export/realisasi-bulanan')
def export_realisasi_bulanan():
    """Export Rekapitulasi Realisasi Bulanan"""
    tahun = request.args.get('tahun')
    bulan = request.args.get('bulan', type=int)
    data = get_realisasi(tahun=tahun, bulan=bulan)
    headers = ['Kode Rekening', 'Uraian', 'Anggaran', 'Realisasi', 'Selisih']
    
    bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    nama_bulan = bulan_nama[bulan-1] if bulan else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Realisasi_Bulanan_{tahun}_{nama_bulan}.xlsx',
        f'REKAPITULASI REALISASI BULAN {nama_bulan} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Realisasi_Bulanan_{tahun}_{nama_bulan}.xlsx')

@app.route('/export/realisasi-tahapan')
def export_realisasi_tahapan():
    """Export Rekapitulasi Realisasi Tahapan"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan')
    data = get_realisasi(tahun=tahun, tahapan=tahapan)
    headers = ['Kode Rekening', 'Uraian', 'Anggaran', 'Realisasi', 'Selisih']
    
    output = export_to_excel(
        data, headers,
        f'Realisasi_Tahapan_{tahapan}_{tahun}.xlsx',
        f'REKAPITULASI REALISASI TAHAPAN {tahapan or "Semua"} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Realisasi_Tahapan_{tahapan}_{tahun}.xlsx')

@app.route('/export/realisasi-tahunan')
def export_realisasi_tahunan():
    """Export Rekapitulasi Realisasi Tahunan"""
    tahun = request.args.get('tahun')
    data = get_realisasi(tahun=tahun)
    headers = ['Kode Rekening', 'Uraian', 'Anggaran', 'Realisasi', 'Selisih']
    
    output = export_to_excel(
        data, headers,
        f'Realisasi_Tahunan_{tahun}.xlsx',
        f'REKAPITULASI REALISASI TAHUNAN TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Realisasi_Tahunan_{tahun}.xlsx')

# ========== REALISASI BARANG HABIS PAKAI ==========
@app.route('/export/realisasi-bhp-bulanan')
def export_realisasi_bhp_bulanan():
    """Export Realisasi Barang Habis Pakai Bulanan"""
    tahun = request.args.get('tahun')
    bulan = request.args.get('bulan', type=int)
    data = get_realisasi_barang_habis(tahun=tahun, bulan=bulan)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    nama_bulan = bulan_nama[bulan-1] if bulan else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'BHP_Bulanan_{tahun}_{nama_bulan}.xlsx',
        f'REALISASI BARANG HABIS PAKAI BULAN {nama_bulan} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BHP_Bulanan_{tahun}_{nama_bulan}.xlsx')

@app.route('/export/realisasi-bhp-tahapan')
def export_realisasi_bhp_tahapan():
    """Export Realisasi Barang Habis Pakai Tahapan"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan')
    data = get_realisasi_barang_habis(tahun=tahun, tahapan=tahapan)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    output = export_to_excel(
        data, headers,
        f'BHP_Tahapan_{tahapan}_{tahun}.xlsx',
        f'REALISASI BARANG HABIS PAKAI TAHAPAN {tahapan or "Semua"} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BHP_Tahapan_{tahapan}_{tahun}.xlsx')

@app.route('/export/realisasi-bhp-tahunan')
def export_realisasi_bhp_tahunan():
    """Export Realisasi Barang Habis Pakai Tahunan"""
    tahun = request.args.get('tahun')
    data = get_realisasi_barang_habis(tahun=tahun)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    output = export_to_excel(
        data, headers,
        f'BHP_Tahunan_{tahun}.xlsx',
        f'REALISASI BARANG HABIS PAKAI TAHUNAN TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BHP_Tahunan_{tahun}.xlsx')

# ========== REALISASI BARANG MODAL/ASET ==========
@app.route('/export/realisasi-modal-bulanan')
def export_realisasi_modal_bulanan():
    """Export Realisasi Barang Modal/Aset Bulanan"""
    tahun = request.args.get('tahun')
    bulan = request.args.get('bulan', type=int)
    data = get_realisasi_barang_modal(tahun=tahun, bulan=bulan)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    nama_bulan = bulan_nama[bulan-1] if bulan else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Barang_Modal_Bulanan_{tahun}_{nama_bulan}.xlsx',
        f'REALISASI BARANG MODAL/ASET BULAN {nama_bulan} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Modal_Bulanan_{tahun}_{nama_bulan}.xlsx')

@app.route('/export/realisasi-modal-tahapan')
def export_realisasi_modal_tahapan():
    """Export Realisasi Barang Modal/Aset Tahapan"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan')
    data = get_realisasi_barang_modal(tahun=tahun, tahapan=tahapan)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    output = export_to_excel(
        data, headers,
        f'Barang_Modal_Tahapan_{tahapan}_{tahun}.xlsx',
        f'REALISASI BARANG MODAL/ASET TAHAPAN {tahapan or "Semua"} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Modal_Tahapan_{tahapan}_{tahun}.xlsx')

@app.route('/export/realisasi-modal-tahunan')
def export_realisasi_modal_tahunan():
    """Export Realisasi Barang Modal/Aset Tahunan"""
    tahun = request.args.get('tahun')
    data = get_realisasi_barang_modal(tahun=tahun)
    headers = ['Kode Rekening', 'Uraian', 'Volume', 'Satuan', 'Harga Satuan', 'Jumlah']
    
    output = export_to_excel(
        data, headers,
        f'Barang_Modal_Tahunan_{tahun}.xlsx',
        f'REALISASI BARANG MODAL/ASET TAHUNAN TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'Modal_Tahunan_{tahun}.xlsx')

# ========== BUKU PEMBANTU RINCIAN OBJEK BELANJA ==========
@app.route('/export/bpobjek-bulanan')
def export_bpobjek_bulanan():
    """Export Buku Pembantu Rincian Objek Belanja Bulanan"""
    tahun = request.args.get('tahun')
    bulan = request.args.get('bulan', type=int)
    data = get_buku_pembantu_objek(tahun=tahun, bulan=bulan)
    headers = ['Kode Rekening', 'Uraian', 'Volume Total', 'Satuan', 'Jumlah Total']
    
    bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
    nama_bulan = bulan_nama[bulan-1] if bulan else "Semua"
    
    output = export_to_excel(
        data, headers,
        f'Buku_Pembantu_Objek_Belanja_Bulanan_{tahun}_{nama_bulan}.xlsx',
        f'BUKU PEMBANTU RINCIAN OBJEK BELANJA BULAN {nama_bulan} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BPObjek_Bulanan_{tahun}_{nama_bulan}.xlsx')

@app.route('/export/bpobjek-tahapan')
def export_bpobjek_tahapan():
    """Export Buku Pembantu Rincian Objek Belanja Tahapan"""
    tahun = request.args.get('tahun')
    tahapan = request.args.get('tahapan')
    data = get_buku_pembantu_objek(tahun=tahun, tahapan=tahapan)
    headers = ['Kode Rekening', 'Uraian', 'Volume Total', 'Satuan', 'Jumlah Total']
    
    output = export_to_excel(
        data, headers,
        f'Buku_Pembantu_Objek_Belanja_Tahapan_{tahapan}_{tahun}.xlsx',
        f'BUKU PEMBANTU RINCIAN OBJEK BELANJA TAHAPAN {tahapan or "Semua"} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BPObjek_Tahapan_{tahapan}_{tahun}.xlsx')

@app.route('/export/bpobjek-tahunan')
def export_bpobjek_tahunan():
    """Export Buku Pembantu Rincian Objek Belanja Tahunan"""
    tahun = request.args.get('tahun')
    data = get_buku_pembantu_objek(tahun=tahun)
    headers = ['Kode Rekening', 'Uraian', 'Volume Total', 'Satuan', 'Jumlah Total']
    
    output = export_to_excel(
        data, headers,
        f'Buku_Pembantu_Objek_Belanja_Tahunan_{tahun}.xlsx',
        f'BUKU PEMBANTU RINCIAN OBJEK BELANJA TAHUNAN TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BPObjek_Tahunan_{tahun}.xlsx')

# ========== LAPORAN BOSP ==========
@app.route('/export/bosp-semester')
def export_bosp_semester():
    """Export Laporan BOSP Semester"""
    tahun = request.args.get('tahun')
    semester = request.args.get('semester')
    data = get_laporan_bosp(tahun=tahun, semester=semester)
    headers = ['Tahun', 'Total Anggaran', 'Total Realisasi', 'Selisih', 'Status']
    
    output = export_to_excel(
        data, headers,
        f'BOSP_Semester_{semester}_{tahun}.xlsx',
        f'LAPORAN REALISASI PENERIMAAN DAN BELANJA DANA BOSP SEMESTER {semester or "Semua"} TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BOSP_Semester_{semester}_{tahun}.xlsx')

@app.route('/export/bosp-tahunan')
def export_bosp_tahunan():
    """Export Laporan BOSP Tahunan"""
    tahun = request.args.get('tahun')
    data = get_laporan_bosp(tahun=tahun)
    headers = ['Tahun', 'Total Anggaran', 'Total Realisasi', 'Selisih', 'Status']
    
    output = export_to_excel(
        data, headers,
        f'BOSP_Tahunan_{tahun}.xlsx',
        f'LAPORAN REALISASI PENERIMAAN DAN BELANJA DANA BOSP TAHUNAN TAHUN {tahun or "Semua"}'
    )
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'BOSP_Tahunan_{tahun}.xlsx')

# === MAIN ===
if __name__ == '__main__':
    print("=" * 60)
    print("📊 ARKASu Data v2.0 - Laporan Keuangan Lengkap")
    print("=" * 60)
    print(f"  Sekolah: {SEKOLAH}")
    print(f"  NPSN: {NPSN}")
    print(f"  Database: {DB_PATH}")
    print("=" * 60)
    print("🌐 Buka browser: http://localhost:5000")
    print("=" * 60)
    
    app.run(debug=True, port=5000)
